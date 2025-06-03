import streamlit as st
import pandas as pd
import unicodedata
from difflib import get_close_matches
from zipfile import ZipFile
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment

st.set_page_config(page_title="Gerador de Planilhas com Cabeçalho", layout="wide")
st.title("Gerador de Planilhas com Cabeçalho para Empresas")

st.markdown(
    """
    1. Faça upload da planilha **“DADOS DAS EMPRESAS.xlsx”** (blocos de 6 linhas por empresa):
       - Coluna B, linha 1 do bloco: **Razão Social**  
       - Coluna B, linha 2 do bloco: **CNPJ**  
       - Coluna B, linha 3 do bloco: **Endereço (completo)**  
       - Coluna B, linha 4 do bloco: **Telefone**  
       - Coluna B, linha 5 do bloco: **E-mail**  

    2. Em seguida, faça upload das planilhas individuais (.xlsx) de cada empresa.  
       O nome do arquivo deve corresponder (ao menos parcialmente) à razão social para que a correspondência funcione.  

    **Observação sobre mesclagem dinâmica do cabeçalho**  
    Cada planilha “empresa.xlsx” possui:
    - O cabeçalho da tabela (“ITEM”, “DESCRIÇÃO…”, etc.) na **linha 6**
    - Os dados reais dos itens a partir da **linha 7**  

    Para determinar até qual coluna mesclar as 5 linhas de cabeçalho, este script:
    1. Examina **linha 6** (os títulos da tabela) e **linha 7** (primeiros dados) antes de inserir cabeçalhos, para descobrir qual é a última coluna não vazia.  
    2. Insere 5 linhas em branco no topo.  
    3. Mescla cada uma das linhas 1 a 5 de **coluna A até essa última coluna detectada**.  
       Ex.: se houver dados só até G6/G7, mescla A1:G1, A2:G2… A5:G5. Se houver até I6/I7, mescla A1:I1, etc.  
    """
)

# 1) Upload da planilha “DADOS DAS EMPRESAS.xlsx”
dados_empresas_file = st.file_uploader(
    "1) Selecione o arquivo DADOS DAS EMPRESAS.xlsx",
    type="xlsx",
    key="dados_empresas"
)

# 2) Upload das planilhas individuais por empresa
arquivos_empresas = st.file_uploader(
    "2) Selecione as planilhas individuais por empresa (.xlsx)",
    type="xlsx",
    accept_multiple_files=True,
    key="arquivos_empresas"
)

def normalizar(texto: str) -> str:
    """
    Remove acentos e caracteres não alfanuméricos, converte para minúsculas.
    """
    nfkd = unicodedata.normalize("NFKD", texto)
    ascii_txt = nfkd.encode("ASCII", "ignore").decode("utf-8")
    return "".join(c for c in ascii_txt if c.isalnum() or c.isspace()).lower().strip()

def extrair_blocos_empresas(df: pd.DataFrame) -> dict:
    """
    Agrupa o DataFrame sem cabeçalho em blocos de 6 linhas cada, extraindo da coluna B:
      - Linha 1 do bloco → Razão Social
      - Linha 2 do bloco → CNPJ
      - Linha 3 do bloco → Endereço completo
      - Linha 4 do bloco → Telefone
      - Linha 5 do bloco → E-mail
    Retorna:
      {
        "Razão Social X": {
            "RAZAO_SOCIAL": ...,
            "CNPJ": ...,
            "ENDERECO": ...,
            "TELEFONE": ...,
            "E-MAIL": ...
        },
        ...
      }
    """
    dados = {}
    for i in range(0, len(df), 6):
        bloco = df.iloc[i : i + 6].reset_index(drop=True)
        if bloco.shape[0] < 2 or pd.isna(bloco.iloc[0, 1]):
            continue
        razao = str(bloco.iloc[0, 1]).strip()
        dados[razao] = {
            "RAZAO_SOCIAL": razao,
            "CNPJ": str(bloco.iloc[1, 1]).strip() if bloco.shape[0] > 1 else "",
            "ENDERECO": str(bloco.iloc[2, 1]).strip() if bloco.shape[0] > 2 else "",
            "TELEFONE": str(bloco.iloc[3, 1]).strip() if bloco.shape[0] > 3 else "",
            "E-MAIL": str(bloco.iloc[4, 1]).strip() if bloco.shape[0] > 4 else "",
        }
    return dados

if dados_empresas_file and arquivos_empresas:
    # 1) Carrega “DADOS DAS EMPRESAS.xlsx” sem cabeçalho
    try:
        df_empresas = pd.read_excel(dados_empresas_file, header=None)
    except Exception as e:
        st.error(f"❌ Erro ao ler “DADOS DAS EMPRESAS.xlsx”: {e}")
        st.stop()

    # 2) Extrai blocos de 6 linhas
    dados_empresas = extrair_blocos_empresas(df_empresas)
    if not dados_empresas:
        st.error("❌ Não foram encontrados blocos válidos em “DADOS DAS EMPRESAS.xlsx”.")
        st.stop()

    # 3) Normaliza chaves (razões sociais) para correspondência aproximada
    dados_empresas_norm = { normalizar(nome): info for nome, info in dados_empresas.items() }

    st.subheader("Razões Sociais Detectadas")
    c1, c2 = st.columns(2)
    for idx, razao in enumerate(dados_empresas.keys()):
        if idx % 2 == 0:
            c1.write(f"- {razao}")
        else:
            c2.write(f"- {razao}")

    # 4) Preparar ZIP em memória
    output_zip = BytesIO()
    match_log = []

    with ZipFile(output_zip, "w") as zipf:
        for arquivo in arquivos_empresas:
            nome_arquivo = arquivo.name
            base_nome = nome_arquivo.replace(".xlsx", "").strip()
            nome_norm = normalizar(base_nome)

            # 4.1) Encontra correspondência aproximada
            matches = get_close_matches(nome_norm, dados_empresas_norm.keys(), n=1, cutoff=0.3)
            if not matches:
                match_log.append(f"❌ NÃO ENCONTRADO: {base_nome} (normalizado: {nome_norm})")
                continue

            chave = matches[0]
            info = dados_empresas_norm[chave]
            match_log.append(f"✅ {base_nome} → {info['RAZAO_SOCIAL']}")

            # 4.2) Abre a planilha individual
            try:
                wb = load_workbook(arquivo)
            except Exception as e:
                match_log.append(f"⚠️ Erro ao abrir “{nome_arquivo}”: {e}")
                continue

            ws = wb.active

            # 4.3) Detecta a última coluna não vazia na linha 6 ORIGINAL:
            #      varre tanto a linha 6 (cabeçalhos da tabela) quanto a linha 7 (primeiros dados),
            #      pois às vezes o cabeçalho pode ter células vazias, mas a linha 7 já possui valor.
            last_col = 1
            # Itera sobre valores da linha 6
            row6_vals = next(ws.iter_rows(min_row=6, max_row=6, values_only=True))
            for idx_col, val in enumerate(row6_vals, start=1):
                if val is not None and str(val).strip() != "":
                    last_col = idx_col
            # Itera sobre valores da linha 7 (caso algum cabeçalho estivesse vazio mas a célula de dados em 7 tenha valor)
            row7_vals = next(ws.iter_rows(min_row=7, max_row=7, values_only=True))
            for idx_col, val in enumerate(row7_vals, start=1):
                if val is not None and str(val).strip() != "":
                    last_col = max(last_col, idx_col)

            # 4.4) Insere 5 linhas vazias no topo (para as 5 linhas de cabeçalho)
            ws.insert_rows(1, amount=5)

            # 4.5) Mescla e preenche cada linha de 1 a 5 até a coluna final detectada
            # LINHA 1 (Razão Social)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
            c1 = ws.cell(row=1, column=1)
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c1.value = f"RAZÃO SOCIAL: {info['RAZAO_SOCIAL']}"

            # LINHA 2 (CNPJ)
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
            c2 = ws.cell(row=2, column=1)
            c2.alignment = Alignment(horizontal="left", vertical="center")
            c2.value = f"CNPJ: {info['CNPJ']}"

            # LINHA 3 (Endereço – com wrap_text)
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
            c3 = ws.cell(row=3, column=1)
            c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            c3.value = f"ENDEREÇO: {info['ENDERECO']}"

            # LINHA 4 (Telefone)
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=last_col)
            c4 = ws.cell(row=4, column=1)
            c4.alignment = Alignment(horizontal="left", vertical="center")
            c4.value = f"TELEFONE: {info['TELEFONE']}"

            # LINHA 5 (E-mail)
            ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=last_col)
            c5 = ws.cell(row=5, column=1)
            c5.alignment = Alignment(horizontal="left", vertical="center")
            c5.value = f"E-MAIL: {info['E-MAIL']}"

            # 4.6) Salvar a planilha modificada em memória e adicionar ao ZIP
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            zipf.writestr(nome_arquivo, buffer.read())

    # 5) Exibe log de correspondências
    st.subheader("Log de Correspondências")
    for linha in match_log:
        st.write(linha)

    # 6) Botão de download do ZIP final
    output_zip.seek(0)
    st.download_button(
        label="📥 Baixar ZIP com Planilhas Formatadas",
        data=output_zip.getvalue(),
        file_name="Planilhas_Cabecalho_Formatado.zip",
        mime="application/zip"
    )

else:
    st.info("Aguardando upload:\n\n1. DADOS DAS EMPRESAS.xlsx\n2. Planilhas por empresa (.xlsx)")
